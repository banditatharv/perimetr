#!/usr/bin/env python3
"""
nmap_to_excel.py

Parses <ip>.nmap files listed in an ips.txt and writes an Excel report
with merged & centered IP cells (or a CSV fallback).

Usage:
    python3 nmap_to_excel.py               # prompts for ips file
    python3 nmap_to_excel.py --ips ips.txt
    python3 nmap_to_excel.py --ips ips.txt --out my_report.xlsx
    python3 nmap_to_excel.py --ips ips.txt --csv-only

Requirements:
    - Python 3.7+
    - openpyxl (optional, if absent script writes CSV only)
"""

import argparse
from pathlib import Path
import re
import csv
from itertools import groupby

import perimetrUI as ui

def parse_nmap_file(path, ip):
    rows = []
    if not path.exists():
        return rows
    with path.open(errors="ignore", encoding="utf-8") as fh:
        for ln in fh:
            ln = ln.rstrip("\n")
            if "/tcp" not in ln:
                continue
            if "TRACEROUTE" in ln:
                continue
            # Normalize whitespace and split into up to 4 parts:
            parts = re.split(r"\s+", ln.strip(), maxsplit=3)
            port_proto = parts[0] if len(parts) > 0 else ""
            state = parts[1] if len(parts) > 1 else ""
            service = parts[2] if len(parts) > 2 else ""
            info = parts[3] if len(parts) > 3 else ""
            if "/" in port_proto:
                port, proto = port_proto.split("/", 1)
            else:
                port, proto = port_proto, ""
            rows.append({
                "ip": ip,
                "port": port,
                "proto": proto,
                "state": state,
                "service": service,
                "info": info
            })
    return rows

def write_csv(rows, outpath):
    with outpath.open("w", newline="", encoding="utf-8") as cf:
        writer = csv.writer(cf)
        writer.writerow(["ip","port","proto","state","service","info"])
        for ip, grp in groupby(rows, key=lambda r: r["ip"]):
            first = True
            for r in grp:
                ip_cell = ip if first else ""
                # Keep info raw; CSV writer will quote if needed
                writer.writerow([ip_cell, r["port"], r["proto"], r["state"], r["service"], r["info"]])
                first = False

def write_xlsx(rows, outpath):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except Exception as e:
        raise RuntimeError("openpyxl not available") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Nmap Report"
    header = ["ip","port","proto","state","service","info"]
    ws.append(header)

    # write grouped rows, remember start/end to merge IP column
    for ip, grp in groupby(rows, key=lambda r: r["ip"]):
        group_list = list(grp)
        start_row = ws.max_row + 1
        for r in group_list:
            ws.append([r["ip"], r["port"], r["proto"], r["state"], r["service"], r["info"]])
        end_row = ws.max_row
        if end_row - start_row + 1 > 1:
            merge_range = f"A{start_row}:A{end_row}"
            ws.merge_cells(merge_range)
            cell = ws[f"A{start_row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
        else:
            cell = ws[f"A{start_row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    # Simple auto-width heuristic
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except:
                val = ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

    wb.save(outpath)

def main():
    p = argparse.ArgumentParser(description="Parse <ip>.nmap files and create merged IP Excel/CSV.")
    p.add_argument("--ips", "-i", type=Path, help="Path to ips.txt (one IP per line). If omitted you will be prompted.")
    p.add_argument("--out", "-o", type=Path, default=Path("nmap_report.xlsx"), help="Output xlsx path (default nmap_report.xlsx). If csv-only or openpyxl missing, a CSV will be written instead.")
    p.add_argument("--csv-only", action="store_true", help="Always write CSV instead of XLSX.")
    args = p.parse_args()

    ui.banner("Service Scan Report Generator", "Merges <ip>.nmap files into one Excel/CSV report")

    ips_path = args.ips
    if not ips_path:
        # interactive prompt
        v = input("Enter path to ips.txt (one IP per line) [ips.txt]: ").strip()
        ips_path = Path(v or "ips.txt")

    if not ips_path.exists():
        ui.error(f"{ips_path} does not exist. Exiting.")
        return

    # read IPs
    ips = []
    with ips_path.open(encoding="utf-8") as fh:
        for l in fh:
            s = l.strip()
            if s:
                ips.append(s)

    if not ips:
        ui.error("No IPs found in ips file. Exiting.")
        return

    ui.info(f"Loaded {len(ips)} IP(s) from {ips_path}")

    # parse each <ip>.nmap
    all_rows = []
    for ip in ips:
        nmap_file = Path(f"{ip}.nmap")
        parsed = parse_nmap_file(nmap_file, ip)
        if parsed:
            all_rows.extend(parsed)
        else:
            # placeholder row so IP still appears in output
            all_rows.append({"ip": ip, "port":"", "proto":"", "state":"", "service":"", "info":"(no results or file missing)"})

    if not all_rows:
        ui.error("No matching /tcp lines found for any IP. Exiting.")
        return

    # If user asked csv-only, or openpyxl missing, write CSV
    if args.csv_only:
        csv_out = args.out.with_suffix(".csv")
        write_csv(all_rows, csv_out)
        ui.success(f"Wrote CSV: {csv_out.resolve()}")
        return

    try:
        write_xlsx(all_rows, args.out)
        ui.success(f"Wrote XLSX: {args.out.resolve()}")
    except Exception as e:
        # fallback to CSV
        csv_out = args.out.with_suffix(".csv")
        write_csv(all_rows, csv_out)
        ui.warn(f"Failed to write XLSX (openpyxl missing or error): {e}")
        ui.success(f"Wrote CSV instead: {csv_out.resolve()}")

if __name__ == "__main__":
    main()